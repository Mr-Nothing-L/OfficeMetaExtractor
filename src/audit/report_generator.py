"""Audit report generation (summary and detail tables)."""
from collections import defaultdict
from datetime import datetime
from typing import List, Dict, Any

from ..utils.datamodel import DocumentMeta, AuditAlert, CompanyAuditResult

# Excel styling constants (kept here to avoid a UI dependency)
_EXCEL_HEADER_FILL = "B88C28"
_EXCEL_HEADER_FONT = "FFFFFF"


def _fmt_range(time_range):
    """Format a (min, max) datetime tuple."""
    if not time_range:
        return ''
    start, end = time_range
    if start == end:
        return start.isoformat()
    return f'{start.isoformat()} ~ {end.isoformat()}'


def _build_company_results(results: List[DocumentMeta]) -> List[CompanyAuditResult]:
    """Aggregate DocumentMeta results into per-company audit results."""
    groups = defaultdict(lambda: {
        'files': [],
        'authors': set(),
        'modifiers': set(),
        'created': [],
        'modified': [],
        'templates': set(),
    })

    for meta in results:
        company = (meta.company or '').strip() or '未知公司'
        g = groups[company]
        g['files'].append(meta.filepath)
        if meta.author:
            g['authors'].add(meta.author.strip())
        if meta.last_modified_by:
            g['modifiers'].add(meta.last_modified_by.strip())
        if isinstance(meta.created, datetime):
            g['created'].append(meta.created)
        if isinstance(meta.modified, datetime):
            g['modified'].append(meta.modified)
        if meta.template:
            g['templates'].add(meta.template.strip())

    company_results = []
    for company, g in sorted(groups.items()):
        creation_range = None
        if g['created']:
            creation_range = (min(g['created']), max(g['created']))
        modification_range = None
        if g['modified']:
            modification_range = (min(g['modified']), max(g['modified']))

        company_results.append(CompanyAuditResult(
            company_name=company,
            file_count=len(g['files']),
            authors=sorted(g['authors']),
            last_modified_by=sorted(g['modifiers']),
            creation_time_range=creation_range,
            modification_time_range=modification_range,
            templates_used=sorted(g['templates']),
        ))

    return company_results


def generate_summary_table(results: List[DocumentMeta],
                           alerts: List[AuditAlert]) -> List[Dict[str, Any]]:
    """Generate company-level summary rows for the audit report."""
    company_results = _build_company_results(results)
    company_results = sorted(company_results, key=lambda x: x.company_name)

    rows = []
    for cr in company_results:
        rows.append({
            '公司名称': cr.company_name,
            '文件数量': cr.file_count,
            '作者列表': ', '.join(cr.authors),
            '最后修改者列表': ', '.join(cr.last_modified_by),
            '创建时间范围': _fmt_range(cr.creation_time_range),
            '修改时间范围': _fmt_range(cr.modification_time_range),
            '使用模板': ', '.join(cr.templates_used),
        })
    return rows


def generate_detail_table(results: List[DocumentMeta],
                          alerts: List[AuditAlert]) -> List[Dict[str, Any]]:
    """Generate detailed finding rows for the audit report."""
    sorted_alerts = sorted(alerts, key=lambda a: (a.rule_name, a.description))

    rows = []
    for alert in sorted_alerts:
        rows.append({
            '规则名称': alert.rule_name,
            '描述': alert.description,
            '涉及公司': ', '.join(alert.affected_companies),
            '涉及文件数': len(alert.affected_files),
            '判定依据': str(alert.details),
        })
    return rows


def export_to_excel(summary_data: List[Dict[str, Any]],
                    detail_data: List[Dict[str, Any]],
                    output_path: str) -> bool:
    """Export summary and detail tables to a multi-sheet Excel file."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill

        wb = Workbook()

        def write_sheet(ws, title, rows):
            ws.title = title
            if not rows:
                return
            headers = list(rows[0].keys())

            header_font = Font(bold=True, color=_EXCEL_HEADER_FONT)
            header_fill = PatternFill(start_color=_EXCEL_HEADER_FILL, end_color=_EXCEL_HEADER_FILL, fill_type='solid')
            header_align = Alignment(horizontal='center', vertical='center')

            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align

            for row_idx, row in enumerate(rows, 2):
                for col_idx, header in enumerate(headers, 1):
                    value = row.get(header, '')
                    ws.cell(row=row_idx, column=col_idx, value=value)

            def _visual_width(text):
                """Approximate visual width: ASCII=1, CJK/other wide chars=2."""
                width = 0
                for ch in str(text):
                    width += 2 if ord(ch) > 127 else 1
                return width

            # Auto column widths (CJK characters count as ~2 widths)
            for col in ws.columns:
                max_width = 0
                col_letter = col[0].column_letter
                for cell in col:
                    try:
                        if cell.value is not None:
                            max_width = max(max_width, _visual_width(cell.value))
                    except Exception:
                        pass
                ws.column_dimensions[col_letter].width = max(min(max_width + 3, 60), 10)

        write_sheet(wb.active, '公司汇总', summary_data)

        detail_ws = wb.create_sheet('详细发现')
        write_sheet(detail_ws, '详细发现', detail_data)

        wb.save(output_path)
        return True
    except Exception as e:
        from ..utils.logger import logger
        logger.error(f'Failed to export audit report: {e}')
        return False
