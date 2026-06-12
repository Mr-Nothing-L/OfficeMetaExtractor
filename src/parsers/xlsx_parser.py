"""XLSX parser implementation."""
from pathlib import Path
from datetime import datetime

from openpyxl import load_workbook

from .base_parser import BaseParser
from ..utils.datamodel import DocumentMeta


class XlsxParser(BaseParser):
    SUPPORTED_EXTENSIONS = ['.xlsx']
    
    # XLSX 文件头: PK (zip格式)
    HEADER = [b'PK\x03\x04', b'PK\x05\x06']
    
    @classmethod
    def _validate_header(cls, filepath: Path) -> bool:
        return cls._check_header(filepath, cls.HEADER)
    
    def parse(self, filepath: Path) -> DocumentMeta:
        meta = self._make_meta(filepath, 'XLSX')
        
        try:
            from openpyxl import load_workbook
            # 尝试正常读取（不用 read_only，避免颜色格式兼容问题）
            wb = load_workbook(filepath)
            cp = wb.properties
            
            meta.author = self._safe_str(cp.creator)
            meta.last_modified_by = self._safe_str(cp.last_modified_by)
            meta.title = self._safe_str(cp.title)
            meta.subject = self._safe_str(cp.subject)
            meta.keywords = self._safe_str(cp.keywords)
            meta.comments = self._safe_str(cp.description)
            meta.revision = self._safe_str(cp.revision)
            
            if cp.created:
                meta.created = cp.created
            if cp.modified:
                meta.modified = cp.modified
            
            wb.close()
            
        except BaseException as e:
            # 如果 openpyxl 读取失败，尝试用 zipfile 提取基础元信息
            try:
                self._parse_fallback(filepath, meta)
                # 备用方案能走到这里就算成功，即使 core.xml 为空
                meta.parse_success = True
                if meta.author or meta.last_modified_by or meta.title:
                    meta.error_message = f"openpyxl 解析失败({type(e).__name__}: {str(e)})，已使用备用方案提取部分信息"
                else:
                    meta.error_message = f"openpyxl 解析失败({type(e).__name__}: {str(e)})，文件core.xml中无元信息"
            except BaseException as fallback_error:
                meta.parse_success = False
                meta.error_message = f"XLSX解析失败: {type(e).__name__}: {str(e)}; 备用方案也失败: {type(fallback_error).__name__}: {str(fallback_error)}"
        
        return meta
    
    def _parse_fallback(self, filepath: Path, meta: DocumentMeta) -> None:
        """当 openpyxl 无法读取时，用 zipfile 直接提取 docProps/core.xml 的基础信息"""
        import zipfile
        from xml.etree import ElementTree as ET
        
        ns = {
            'cp': 'http://schemas.openxmlformats.org/package/2006/metadata/core-properties',
            'dc': 'http://purl.org/dc/elements/1.1/',
            'dcterms': 'http://purl.org/dc/terms/',
        }
        
        with zipfile.ZipFile(filepath, 'r') as zf:
            # 尝试读取核心属性
            if 'docProps/core.xml' in zf.namelist():
                with zf.open('docProps/core.xml') as f:
                    tree = ET.parse(f)
                    root = tree.getroot()
                    
                    # 提取基础字段
                    creator = root.find('dc:creator', ns)
                    if creator is not None and creator.text:
                        meta.author = creator.text
                    
                    last_mod = root.find('cp:lastModifiedBy', ns)
                    if last_mod is not None and last_mod.text:
                        meta.last_modified_by = last_mod.text
                    
                    title = root.find('dc:title', ns)
                    if title is not None and title.text:
                        meta.title = title.text
                    
                    subject = root.find('dc:subject', ns)
                    if subject is not None and subject.text:
                        meta.subject = subject.text
                    
                    # 时间字段
                    created = root.find('dcterms:created', ns)
                    if created is not None and created.text:
                        from datetime import datetime
                        try:
                            meta.created = datetime.fromisoformat(created.text.replace('Z', '+00:00'))
                        except:
                            pass
                    
                    modified = root.find('dcterms:modified', ns)
                    if modified is not None and modified.text:
                        from datetime import datetime
                        try:
                            meta.modified = datetime.fromisoformat(modified.text.replace('Z', '+00:00'))
                        except:
                            pass
            
            # 尝试读取扩展属性（公司名等）
            if 'docProps/app.xml' in zf.namelist():
                with zf.open('docProps/app.xml') as f:
                    tree = ET.parse(f)
                    root = tree.getroot()
                    # 这里可以提取更多属性
        
        # 备用方案能走到这里就算成功（即使 core.xml 为空）
        meta.parse_success = True
