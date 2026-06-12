"""Result table widget for displaying extracted metadata."""
from PyQt5.QtWidgets import (
    QTableWidget, QTableWidgetItem, QAbstractItemView, QHeaderView,
    QMenu, QAction, QMessageBox, QFileDialog
)
from PyQt5.QtCore import Qt, pyqtSignal
from PyQt5.QtGui import QColor, QFont
import os
import subprocess
from typing import List, Dict, Any


class ResultTable(QTableWidget):
    """Table displaying document metadata extraction results."""
    
    item_double_clicked = pyqtSignal(str)
    
    COLUMN_HEADERS = [
        "文件名", "格式", "作者", "最后编辑者",
        "创建时间", "修改时间", "标题", "主题", "状态"
    ]
    COLUMN_KEYS = [
        'filename', 'format', 'author', 'last_modified_by',
        'created', 'modified', 'title', 'subject', 'status'
    ]
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self._data: List[Dict[str, Any]] = []
        self._init_ui()
    
    def _init_ui(self):
        self.setColumnCount(len(self.COLUMN_HEADERS))
        self.setHorizontalHeaderLabels(self.COLUMN_HEADERS)
        
        self.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.setSelectionMode(QAbstractItemView.ExtendedSelection)
        self.setAlternatingRowColors(True)
        self.setSortingEnabled(True)
        
        # Column sizing
        header = self.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.Stretch)
        header.setSectionResizeMode(1, QHeaderView.Fixed)
        self.setColumnWidth(1, 50)
        for i in range(2, len(self.COLUMN_HEADERS)):
            header.setSectionResizeMode(i, QHeaderView.Interactive)
        
        self.setContextMenuPolicy(Qt.CustomContextMenu)
        self.customContextMenuRequested.connect(self._show_context_menu)
        self.cellDoubleClicked.connect(self._on_double_click)
    
    def set_data(self, data: List[Dict[str, Any]]):
        """Populate table with data."""
        self._data = data
        self.setSortingEnabled(False)
        self.setRowCount(len(data))
        
        for row_idx, item in enumerate(data):
            success = not str(item.get('status', '')).startswith('失败')
            
            for col_idx, key in enumerate(self.COLUMN_KEYS):
                val = item.get(key, '')
                if val is None:
                    val = ''
                cell = QTableWidgetItem(str(val))
                cell.setFlags(Qt.ItemIsSelectable | Qt.ItemIsEnabled)
                cell.setData(Qt.UserRole, item.get('filepath', ''))
                
                if not success:
                    cell.setForeground(QColor("#f48771"))
                
                self.setItem(row_idx, col_idx, cell)
            
            # Set row height
            self.setRowHeight(row_idx, 24)
        
        self.setSortingEnabled(True)
        self.resizeColumnsToContents()
        # Reset first column to stretch after resize
        self.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
    
    def get_data(self) -> List[Dict[str, Any]]:
        """Return current data."""
        return self._data
    
    def clear_data(self):
        """Clear all data."""
        self._data = []
        self.setRowCount(0)
    
    def selected_rows_data(self) -> List[Dict[str, Any]]:
        """Get data for selected rows."""
        selected = []
        for idx in self.selectionModel().selectedRows():
            row = idx.row()
            if 0 <= row < len(self._data):
                selected.append(self._data[row])
        return selected
    
    def _show_context_menu(self, position):
        menu = QMenu(self)
        
        copy_action = QAction("复制单元格", self)
        copy_action.triggered.connect(self._copy_cell)
        menu.addAction(copy_action)
        
        copy_row_action = QAction("复制整行", self)
        copy_row_action.triggered.connect(self._copy_row)
        menu.addAction(copy_row_action)
        
        menu.addSeparator()
        
        open_loc_action = QAction("打开文件位置", self)
        open_loc_action.triggered.connect(self._open_file_location)
        menu.addAction(open_loc_action)
        
        menu.exec_(self.viewport().mapToGlobal(position))
    
    def _copy_cell(self):
        item = self.currentItem()
        if item:
            from PyQt5.QtWidgets import QApplication
            QApplication.clipboard().setText(item.text())
    
    def _copy_row(self):
        row = self.currentRow()
        if row < 0 or row >= len(self._data):
            return
        
        values = []
        for col in range(self.columnCount()):
            item = self.item(row, col)
            values.append(item.text() if item else '')
        
        from PyQt5.QtWidgets import QApplication
        QApplication.clipboard().setText('\t'.join(values))
    
    def _open_file_location(self):
        row = self.currentRow()
        if row < 0 or row >= len(self._data):
            return
        
        filepath = self._data[row].get('filepath', '')
        if not filepath or not os.path.exists(filepath):
            QMessageBox.warning(self, "错误", "文件不存在")
            return
        
        directory = os.path.dirname(filepath)
        if sys.platform == 'win32':
            subprocess.run(['explorer', '/select,', filepath])
        elif sys.platform == 'darwin':
            subprocess.run(['open', directory])
        else:
            subprocess.run(['xdg-open', directory])
    
    def _on_double_click(self, row, col):
        if row < 0 or row >= len(self._data):
            return
        filepath = self._data[row].get('filepath', '')
        if filepath:
            self.item_double_clicked.emit(filepath)
    
    def export_csv(self, filepath: str) -> bool:
        """Export data to CSV."""
        import csv
        try:
            with open(filepath, 'w', newline='', encoding='utf-8-sig') as f:
                writer = csv.writer(f)
                writer.writerow(self.COLUMN_HEADERS)
                for item in self._data:
                    writer.writerow([item.get(k, '') for k in self.COLUMN_KEYS])
            return True
        except Exception as e:
            QMessageBox.critical(self, "导出失败", f"导出 CSV 失败:\n{str(e)}")
            return False
    
    def export_json(self, filepath: str) -> bool:
        """Export data to JSON."""
        import json
        try:
            with open(filepath, 'w', encoding='utf-8') as f:
                json.dump(self._data, f, ensure_ascii=False, indent=2)
            return True
        except Exception as e:
            QMessageBox.critical(self, "导出失败", f"导出 JSON 失败:\n{str(e)}")
            return False
    
    def export_excel(self, filepath: str) -> bool:
        """Export data to Excel."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Metadata"
            
            # Header
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="0E639C", end_color="0E639C", fill_type="solid")
            header_align = Alignment(horizontal="center", vertical="center")
            
            for col_idx, header in enumerate(self.COLUMN_HEADERS, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
            
            # Data
            for row_idx, item in enumerate(self._data, 2):
                for col_idx, key in enumerate(self.COLUMN_KEYS, 1):
                    val = item.get(key, '')
                    if val is None:
                        val = ''
                    cell = ws.cell(row=row_idx, column=col_idx, value=str(val))
                    if not str(item.get('status', '')).startswith('失败'):
                        cell.font = Font(color="F48771")
            
            # Auto column widths
            for col in ws.columns:
                max_len = 0
                col_letter = col[0].column_letter
                for cell in col:
                    try:
                        if cell.value:
                            max_len = max(max_len, len(str(cell.value)))
                    except Exception:
                        pass
                adjusted = min(max_len + 2, 50)
                ws.column_dimensions[col_letter].width = adjusted
            
            wb.save(filepath)
            return True
            
        except Exception as e:
            QMessageBox.critical(self, "导出失败", f"导出 Excel 失败:\n{str(e)}")
            return False


import sys
