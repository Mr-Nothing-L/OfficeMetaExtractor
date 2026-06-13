"""Audit result table widget using QTableView + virtual model."""
import csv
import json
import os
import subprocess
import sys
from typing import Any, Dict, List

from PyQt5.QtWidgets import (
    QTableView, QAbstractItemView, QHeaderView, QMenu, QAction,
    QMessageBox, QFileDialog
)
from PyQt5.QtCore import Qt, pyqtSignal, QModelIndex
from PyQt5.QtGui import QColor

from .styles import Theme
from .models import AuditTableModel


class AuditResultTable(QTableView):
    """Table displaying batch/audit extraction results."""

    item_double_clicked = pyqtSignal(str)
    show_audit_detail = pyqtSignal(dict)

    def __init__(self, parent=None):
        super().__init__(parent)
        self._model = AuditTableModel(self)
        self.setModel(self._model)
        self._init_ui()

    def _init_ui(self):
        self.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.setSelectionMode(QAbstractItemView.ExtendedSelection)
        self.setAlternatingRowColors(True)
        self.setSortingEnabled(True)

        header = self.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.Stretch)
        header.setSectionResizeMode(1, QHeaderView.Fixed)
        self.setColumnWidth(1, 50)
        header.setSectionResizeMode(2, QHeaderView.Interactive)  # 公司
        for i in range(3, len(self._model.COLUMN_HEADERS)):
            header.setSectionResizeMode(i, QHeaderView.Interactive)

        self.setContextMenuPolicy(Qt.CustomContextMenu)
        self.customContextMenuRequested.connect(self._show_context_menu)
        self.doubleClicked.connect(self._on_double_click)

    # ------------------------------------------------------------------
    # Data API
    # ------------------------------------------------------------------
    def set_data(self, data: List[Dict[str, Any]]) -> None:
        """Populate table with data."""
        self._model.set_data(data)

    def append_row(self, row: Dict[str, Any]) -> None:
        """Append a single row to the table."""
        self._model.append_row(row)

    def append_rows(self, rows: List[Dict[str, Any]]) -> None:
        """Append multiple rows to the table."""
        self._model.append_rows(rows)

    def get_data(self) -> List[Dict[str, Any]]:
        """Return current data."""
        return self._model.get_data()

    def clear_data(self) -> None:
        """Clear all data."""
        self._model.clear()

    def selected_rows_data(self) -> List[Dict[str, Any]]:
        """Get data for selected rows."""
        selected = []
        for idx in self.selectionModel().selectedRows():
            row = idx.row()
            if 0 <= row < len(self._model.get_data()):
                selected.append(self._model.get_data()[row])
        return selected

    # ------------------------------------------------------------------
    # Interactions
    # ------------------------------------------------------------------
    def _show_context_menu(self, position):
        menu = QMenu(self)

        copy_action = QAction("复制单元格", self)
        copy_action.triggered.connect(self._copy_cell)
        menu.addAction(copy_action)

        copy_row_action = QAction("复制整行", self)
        copy_row_action.triggered.connect(self._copy_row)
        menu.addAction(copy_row_action)

        menu.addSeparator()

        open_loc_action = QAction("打开文件位置", self)
        open_loc_action.triggered.connect(self._open_file_location)
        menu.addAction(open_loc_action)

        menu.addSeparator()

        view_detail_action = QAction("查看审计详情", self)
        view_detail_action.triggered.connect(self._on_view_detail)
        menu.addAction(view_detail_action)

        menu.exec_(self.viewport().mapToGlobal(position))

    def _copy_cell(self):
        index = self.currentIndex()
        if index.isValid():
            from PyQt5.QtWidgets import QApplication
            QApplication.clipboard().setText(str(self._model.data(index, Qt.DisplayRole)))

    def _copy_row(self):
        row = self.currentIndex().row()
        if row < 0 or row >= len(self._model.get_data()):
            return
        values = []
        for col in range(self._model.columnCount()):
            idx = self._model.index(row, col)
            values.append(str(self._model.data(idx, Qt.DisplayRole)))
        from PyQt5.QtWidgets import QApplication
        QApplication.clipboard().setText('\t'.join(values))

    def _open_file_location(self):
        row = self.currentIndex().row()
        if row < 0 or row >= len(self._model.get_data()):
            return
        filepath = self._model.get_data()[row].get('filepath', '')
        if not filepath or not os.path.exists(filepath):
            QMessageBox.warning(self, "错误", "文件不存在")
            return
        directory = os.path.dirname(filepath)
        if sys.platform == 'win32':
            subprocess.run(['explorer', '/select,', filepath])
        elif sys.platform == 'darwin':
            subprocess.run(['open', directory])
        else:
            subprocess.run(['xdg-open', directory])

    def _on_view_detail(self):
        """Emit signal to show audit detail for selected row(s)."""
        selected = self.selected_rows_data()
        if not selected:
            return
        row_data = selected[0]
        alert = {
            'rule_name': 'file_audit_summary',
            'severity': 'low',
            'description': (
                f"文件: {row_data.get('filename', '')}\n"
                f"公司: {row_data.get('company', '')}\n"
                f"作者: {row_data.get('author', '')}\n"
                f"最后编辑者: {row_data.get('last_modified_by', '')}"
            ),
            'affected_companies': [row_data.get('company', '')] if row_data.get('company') else [],
            'affected_files': [row_data.get('filepath', '')] if row_data.get('filepath') else [],
            'details': row_data,
        }
        self.show_audit_detail.emit(alert)

    def _on_double_click(self, index: QModelIndex):
        if not index.isValid():
            return
        row = index.row()
        if row < 0 or row >= len(self._model.get_data()):
            return
        filepath = self._model.get_data()[row].get('filepath', '')
        if filepath:
            self.item_double_clicked.emit(filepath)

    # ------------------------------------------------------------------
    # Export
    # ------------------------------------------------------------------
    def export_csv(self, filepath: str) -> bool:
        """Export data to CSV."""
        try:
            data = self._model.get_data()
            with open(filepath, 'w', newline='', encoding='utf-8-sig') as f:
                writer = csv.writer(f)
                writer.writerow(self._model.COLUMN_HEADERS)
                for item in data:
                    writer.writerow([item.get(k, '') for k in self._model.COLUMN_KEYS])
            return True
        except Exception as e:
            QMessageBox.critical(self, "导出失败", f"导出 CSV 失败:\n{str(e)}")
            return False

    def export_json(self, filepath: str) -> bool:
        """Export data to JSON."""
        try:
            with open(filepath, 'w', encoding='utf-8') as f:
                json.dump(self._model.get_data(), f, ensure_ascii=False, indent=2)
            return True
        except Exception as e:
            QMessageBox.critical(self, "导出失败", f"导出 JSON 失败:\n{str(e)}")
            return False

    def export_excel(self, filepath: str) -> bool:
        """Export data to Excel."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill

            wb = Workbook()
            ws = wb.active
            ws.title = "Metadata"

            header_font = Font(bold=True, color=Theme.EXCEL_HEADER_FONT)
            header_fill = PatternFill(
                start_color=Theme.EXCEL_HEADER_FILL,
                end_color=Theme.EXCEL_HEADER_FILL,
                fill_type="solid"
            )
            header_align = Alignment(horizontal="center", vertical="center")

            for col_idx, header in enumerate(self._model.COLUMN_HEADERS, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align

            for row_idx, item in enumerate(self._model.get_data(), 2):
                for col_idx, key in enumerate(self._model.COLUMN_KEYS, 1):
                    val = item.get(key, '')
                    if val is None:
                        val = ''
                    cell = ws.cell(row=row_idx, column=col_idx, value=str(val))
                    if str(item.get('status', '')).startswith('失败'):
                        cell.font = Font(color="F48771")

            for col in ws.columns:
                max_len = 0
                col_letter = col[0].column_letter
                for cell in col:
                    try:
                        if cell.value:
                            max_len = max(max_len, len(str(cell.value)))
                    except Exception:
                        pass
                adjusted = min(max_len + 2, 50)
                ws.column_dimensions[col_letter].width = adjusted

            wb.save(filepath)
            return True

        except Exception as e:
            QMessageBox.critical(self, "导出失败", f"导出 Excel 失败:\n{str(e)}")
            return False
